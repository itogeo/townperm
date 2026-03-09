#!/usr/bin/env python3
"""
Ito Geospatial - Sales Prospect List Generator
Creates a comprehensive outreach list and tracking spreadsheet
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = Workbook()

# ============================================================================
# SHEET 1: A-LIST PROSPECTS (50 High-Priority, Personalized Outreach)
# ============================================================================
ws1 = wb.active
ws1.title = "A-List (50 Priority)"

# Headers
headers_a = [
    "Tier", "State", "Town", "Population", "County", 
    "Contact Name", "Title", "Email", "Phone", "Website",
    "Current System", "Pain Point", "Personalization Hook",
    "Status", "Date Contacted", "Response", "Follow-up Date", "Notes"
]

# Style headers
header_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers_a, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# A-List Data - Montana Towns (Your Home Turf)
a_list_montana = [
    ("A", "MT", "Three Forks", 1919, "Gallatin", "Crystal Turner", "City Clerk", "clerk@threeforksmt.gov", "(406) 285-3431", "https://threeforksmt.gov", "Paper/PDF", "Permit tracking, council reporting", "Already demoed HometownMap - warm lead"),
    ("A", "MT", "Manhattan", 2288, "Gallatin", "Brandon Yung", "Building Official", "planning@manhattanmt.gov", "(406) 595-9300", "https://manhattanmt.gov", "Paper", "No parcel integration", "Neighbor to Three Forks - easy reference"),
    ("A", "MT", "Belgrade", 11872, "Gallatin", "Planning Dept", "City Planner", "planning@cityofbelgrade.net", "(406) 388-3760", "https://cityofbelgrade.net", "Unknown", "Rapid growth, likely paper", "Fastest growing MT city - needs systems"),
    ("A", "MT", "Livingston", 8780, "Park", "City Clerk", "City Clerk", "cityclerk@livingstonmontana.org", "(406) 222-2025", "https://livingstonmontana.org", "Unknown", "Gateway to Yellowstone", "Tourism growth driving permits"),
    ("A", "MT", "Big Timber", 1611, "Sweet Grass", "City Clerk", "Clerk/Treasurer", "bigtimber@itstriangle.com", "(406) 932-5245", "", "Paper", "Small staff, manual processes", "Classic small MT town"),
    ("A", "MT", "Townsend", 1967, "Broadwater", "City Clerk", "Clerk", "clerk@townsendmt.com", "(406) 266-3911", "https://townsendmt.com", "Paper", "Growing, needs efficiency", "Near Helena, spillover growth"),
    ("A", "MT", "Dillon", 4085, "Beaverhead", "Planning Dept", "Planner", "planning@dillonmt.org", "(406) 683-4245", "https://dillonmt.org", "Unknown", "University town, steady permits", "College town dynamics"),
    ("A", "MT", "Red Lodge", 2502, "Carbon", "City Admin", "Administrator", "admin@cityofredlodge.net", "(406) 446-1606", "https://cityofredlodge.net", "Paper", "Tourism, historic district", "Ski town, design review needed"),
    ("A", "MT", "Columbus", 2003, "Stillwater", "City Clerk", "Clerk", "columbus@stillwater.mt.gov", "(406) 322-5313", "", "Paper", "Basic needs", "Growing bedroom community"),
    ("A", "MT", "Ennis", 777, "Madison", "Town Clerk", "Clerk", "townofennis@3rivers.net", "(406) 682-4287", "https://ennismontana.org", "Paper", "Fly fishing tourism", "Small but affluent tourism"),
]

# A-List Data - Maine Towns (We researched Pownal doing 176 permits on paper!)
a_list_maine = [
    ("A", "ME", "Pownal", 1550, "Cumberland", "Millard Billings", "CEO", "ceo@pownalmaine.org", "(207) 688-4611", "https://pownalmaine.org", "Paper/PDF", "176 permits/year on paper!", "CONFIRMED paper - Annual report shows pain"),
    ("A", "ME", "Franklin", 1575, "Hancock", "Millard Billings", "CEO", "ceo@franklinmaine.com", "(207) 565-8806", "https://franklinmaine.com", "Paper", "Part-time CEO, manual tracking", "Only available Mon 9-11am - understaffed"),
    ("A", "ME", "Southwest Harbor", 1764, "Hancock", "Jarrod Kushla", "CEO", "ceo@southwestharbor.org", "(207) 244-5404", "https://southwestharbormaine.org", "Paper", "Tourism area, shoreland permits", "Acadia gateway - complex permits"),
    ("A", "ME", "Bucksport", 4830, "Hancock", "Luke Chiavelli", "CEO", "ceo@bucksportmaine.gov", "(207) 469-7368", "https://bucksportmaine.gov", "Paper", "One person does everything", "CEO also does plumbing, addressing"),
    ("A", "ME", "Sweden", 350, "Oxford", "Peter Gordon", "CEO", "ceo@swedenmaine.org", "(207) 298-6248", "https://swedenmaine.org", "Paper", "Tiny town, minimal hours", "Fridays 11-3 only - needs efficiency"),
    ("A", "ME", "Hermon", 6036, "Penobscot", "Josh Murphy", "CEO", "murphyj@hermonmaine.gov", "(207) 848-1042", "https://hermonmaine.gov", "Unknown", "Growing suburb, PDF forms", "Active planning board"),
    ("A", "ME", "Clinton", 3200, "Kennebec", "CEO", "ceo@clinton-me.us", "(207) 426-8511", "https://clinton-me.us", "Paper", "Multi-role position", "CEO does 6+ different jobs"),
    ("A", "ME", "Monmouth", 4100, "Kennebec", "CEO", "ceo@monmouthmaine.gov", "(207) 933-2206", "https://monmouthmaine.gov", "Paper", "3rd party inspectors used", "Needs coordination tools"),
]

# A-List Data - Wyoming Towns
a_list_wyoming = [
    ("A", "WY", "Dubois", 971, "Fremont", "Town Clerk", "Clerk", "clerk@townofdubois.org", "(307) 455-2345", "", "Paper", "Gateway to wilderness", "Tourism permits"),
    ("A", "WY", "Saratoga", 1690, "Carbon", "Town Clerk", "Clerk", "clerk@saratoga-wy.gov", "(307) 326-8335", "https://saratoga-wy.gov", "Paper", "Hot springs tourism", "Small but active"),
    ("A", "WY", "Pinedale", 1871, "Sublette", "Town Clerk", "Clerk", "clerk@townofpinedale.us", "(307) 367-4136", "https://townofpinedale.us", "Paper", "Energy + tourism", "Boom/bust permit cycles"),
    ("A", "WY", "Lander", 7533, "Fremont", "Planning Dept", "Planner", "planning@landerwyoming.org", "(307) 332-2870", "https://landerwyoming.org", "Unknown", "Outdoor rec hub", "Growing steadily"),
    ("A", "WY", "Thermopolis", 2897, "Hot Springs", "Town Clerk", "Clerk", "clerk@thermopolis.org", "(307) 864-2348", "", "Paper", "Hot springs, small town", "Tourism infrastructure"),
]

# A-List Data - Idaho Towns
a_list_idaho = [
    ("A", "ID", "Idaho City", 485, "Boise", "City Clerk", "Clerk", "clerk@idahocity.gov", "(208) 392-4584", "https://idahocity.municipalimpact.com", "Paper", "Historic mining town", "Website shows PDF permits"),
    ("A", "ID", "Driggs", 1920, "Teton", "City Clerk", "Clerk", "clerk@driggs.id.gov", "(208) 354-2362", "https://driggs.id.gov", "Unknown", "Teton Valley gateway", "Ski area growth"),
    ("A", "ID", "Victor", 2260, "Teton", "City Clerk", "Clerk", "clerk@victorcityidaho.com", "(208) 787-2940", "https://victorcityidaho.com", "Paper", "Adjacent to Jackson Hole", "Spillover development"),
    ("A", "ID", "Salmon", 3112, "Lemhi", "City Clerk", "Clerk", "clerk@cityofsalmonidaho.com", "(208) 756-3214", "", "Paper", "Remote, outdoor rec", "Wilderness permits"),
    ("A", "ID", "McCall", 3286, "Valley", "Planning", "Planner", "planning@mccall.id.us", "(208) 634-3504", "https://mccall.id.us", "Unknown", "Resort town", "High permit volume"),
]

# A-List Data - Other States
a_list_other = [
    ("A", "CO", "Idaho Springs", 1717, "Clear Creek", "City Clerk", "Clerk", "cityclerk@idahospringsco.com", "(303) 567-4421", "https://idahospringsco.com", "Paper", "Mountain town, 2018 IBC", "Website shows manual permits"),
    ("A", "CO", "Silverton", 632, "San Juan", "Town Clerk", "Clerk", "clerk@silverton.co.gov", "(970) 387-5522", "", "Paper", "Historic ski town", "Mining + tourism permits"),
    ("A", "NM", "Red River", 477, "Taos", "Town Clerk", "Clerk", "clerk@redriver.org", "(575) 754-2277", "https://redriver.org", "Paper", "Ski valley", "Seasonal surge"),
    ("A", "NM", "Angel Fire", 1122, "Colfax", "Town Clerk", "Clerk", "clerk@angelfire.org", "(575) 377-3232", "https://angelfire.org", "Paper", "Resort community", "Vacation home permits"),
    ("A", "SD", "Deadwood", 1311, "Lawrence", "City Finance", "Finance Officer", "finance@cityofdeadwood.com", "(605) 578-2082", "https://cityofdeadwood.com", "Paper", "Gaming + historic", "Design review needed"),
    ("A", "SD", "Hill City", 948, "Pennington", "Finance Officer", "Clerk", "clerk@hillcitysd.org", "(605) 574-2300", "https://hillcitysd.org", "Paper", "Rushmore gateway", "Tourism infrastructure"),
    ("A", "NV", "Genoa", 935, "Douglas", "Town Clerk", "Clerk", "clerk@genoanv.org", "(775) 782-8696", "https://genoanv.org", "Paper", "Nevada's oldest town", "Historic preservation"),
    ("A", "AZ", "Jerome", 455, "Yavapai", "Town Clerk", "Clerk", "clerk@jerome.az.gov", "(928) 634-7943", "https://jerome.az.gov", "Paper", "Historic mining town", "Tourism + preservation"),
]

# Combine all A-list
all_a_list = a_list_montana + a_list_maine + a_list_wyoming + a_list_idaho + a_list_other

# Write A-list data
for row, data in enumerate(all_a_list, 2):
    for col, value in enumerate(data, 1):
        cell = ws1.cell(row=row, column=col, value=value)
        cell.border = thin_border
    # Add empty cells for tracking columns
    for col in range(len(data) + 1, len(headers_a) + 1):
        cell = ws1.cell(row=row, column=col, value="")
        cell.border = thin_border
    # Set status to "Not Contacted"
    ws1.cell(row=row, column=14, value="Not Contacted")

# Set column widths
col_widths_a = [5, 5, 18, 10, 12, 18, 15, 30, 15, 35, 12, 25, 40, 15, 12, 20, 12, 40]
for i, width in enumerate(col_widths_a, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# ============================================================================
# SHEET 2: B-LIST PROSPECTS (200 Semi-Personalized)
# ============================================================================
ws2 = wb.create_sheet("B-List (200)")

headers_b = [
    "State", "Town", "Population", "County", "Email Pattern", 
    "Status", "Date Sent", "Opened", "Clicked", "Replied", "Notes"
]

for col, header in enumerate(headers_b, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# B-List Data - Montana towns 1000-5000
b_list_mt = [
    ("MT", "Glendive", 4810, "Dawson", "clerk@cityofglendive.com"),
    ("MT", "Hardin", 3742, "Big Horn", "clerk@hardinmt.com"),
    ("MT", "Shelby", 3216, "Toole", "clerk@shelbymt.com"),
    ("MT", "Glasgow", 3190, "Valley", "clerk@glasgowmt.net"),
    ("MT", "Libby", 3042, "Lincoln", "clerk@cityoflibby.com"),
    ("MT", "Cut Bank", 3028, "Glacier", "clerk@cutbankmt.org"),
    ("MT", "Deer Lodge", 3021, "Powell", "clerk@deerlodgemt.org"),
    ("MT", "Conrad", 2618, "Pondera", "clerk@conradmt.com"),
    ("MT", "Wolf Point", 2602, "Roosevelt", "clerk@wolfpoint.org"),
    ("MT", "Stevensville", 2130, "Ravalli", "clerk@stevensville-mt.org"),
    ("MT", "Colstrip", 2101, "Rosebud", "clerk@colstripmt.org"),
    ("MT", "Malta", 2100, "Phillips", "clerk@maltamt.com"),
    ("MT", "Roundup", 2003, "Musselshell", "clerk@roundupmt.net"),
    ("MT", "Ronan", 2001, "Lake", "clerk@cityofronan.com"),
    ("MT", "Eureka", 1840, "Lincoln", "clerk@eureka-mt.org"),
    ("MT", "Forsyth", 1728, "Rosebud", "clerk@forsythmt.org"),
    ("MT", "Plentywood", 1654, "Sheridan", "clerk@plentywood.com"),
    ("MT", "Choteau", 1479, "Teton", "clerk@choteaumt.org"),
    ("MT", "Fort Benton", 1474, "Chouteau", "clerk@fortbenton.com"),
    ("MT", "West Yellowstone", 1474, "Gallatin", "clerk@townofwestyellowstone.com"),
    ("MT", "Chinook", 1353, "Blaine", "clerk@chinookmt.com"),
    ("MT", "Boulder", 1284, "Jefferson", "clerk@bouldermt.com"),
    ("MT", "Plains", 1195, "Sanders", "clerk@plainsmt.com"),
    ("MT", "Whitehall", 1156, "Jefferson", "clerk@whitehallmt.org"),
    ("MT", "Harlowton", 1058, "Wheatland", "clerk@harlowton.net"),
]

# B-List Data - Maine towns 1000-5000
b_list_me = [
    ("ME", "Searsport", 2722, "Waldo", "clerk@searsport.maine.gov"),
    ("ME", "Winterport", 3757, "Waldo", "clerk@winterportmaine.gov"),
    ("ME", "Stockton Springs", 1591, "Waldo", "clerk@stocktonsprings.org"),
    ("ME", "Unity", 2099, "Waldo", "clerk@unityme.org"),
    ("ME", "Belfast", 6668, "Waldo", "clerk@cityofbelfast.org"),
    ("ME", "Searsmont", 1478, "Waldo", "clerk@searsmont.com"),
    ("ME", "Lincolnville", 2164, "Waldo", "clerk@lincolnvillemaine.org"),
    ("ME", "Northport", 1544, "Waldo", "clerk@northportmaine.org"),
    ("ME", "Islesboro", 566, "Waldo", "clerk@townofislesboro.com"),
    ("ME", "Brooks", 1083, "Waldo", "clerk@brooksme.org"),
    ("ME", "Castine", 1366, "Hancock", "clerk@castine.me.us"),
    ("ME", "Blue Hill", 2686, "Hancock", "clerk@townofbluehill.org"),
    ("ME", "Deer Isle", 1975, "Hancock", "clerk@deerisle.me"),
    ("ME", "Stonington", 1043, "Hancock", "clerk@stonington.me.us"),
    ("ME", "Surry", 1702, "Hancock", "clerk@surrymaine.org"),
    ("ME", "Trenton", 1481, "Hancock", "clerk@trentonme.com"),
    ("ME", "Lamoine", 1720, "Hancock", "clerk@lamoine-me.gov"),
    ("ME", "Gouldsboro", 1737, "Hancock", "clerk@gouldsboromaine.us"),
    ("ME", "Sullivan", 1236, "Hancock", "clerk@sullivanmaine.org"),
    ("ME", "Hancock", 2394, "Hancock", "clerk@townofhancock.us"),
]

# B-List Data - Wyoming
b_list_wy = [
    ("WY", "Powell", 6314, "Park", "clerk@cityofpowell.com"),
    ("WY", "Worland", 4835, "Washakie", "clerk@cityofworland.org"),
    ("WY", "Torrington", 6501, "Goshen", "clerk@torrington-wyoming.com"),
    ("WY", "Newcastle", 3532, "Weston", "clerk@newcastlewyoming.org"),
    ("WY", "Afton", 1911, "Lincoln", "clerk@aftonwyoming.org"),
    ("WY", "Kemmerer", 2656, "Lincoln", "clerk@kemmerer.org"),
    ("WY", "Mountain View", 1286, "Uinta", "clerk@mountainviewwy.gov"),
    ("WY", "Green River", 11808, "Sweetwater", "clerk@cityofgreenriver.org"),
    ("WY", "Evanston", 11879, "Uinta", "clerk@evanstonwy.gov"),
    ("WY", "Douglas", 6386, "Converse", "clerk@cityofdouglas.org"),
    ("WY", "Buffalo", 4585, "Johnson", "clerk@buffalowy.gov"),
    ("WY", "Wheatland", 3627, "Platte", "clerk@wheatlandwyoming.org"),
    ("WY", "Glenrock", 2576, "Converse", "clerk@glenrockwy.org"),
    ("WY", "Lovell", 2360, "Big Horn", "clerk@lovellwy.gov"),
    ("WY", "Basin", 1285, "Big Horn", "clerk@basinwyoming.com"),
]

# B-List Data - Idaho
b_list_id = [
    ("ID", "Sandpoint", 8639, "Bonner", "clerk@sandpointidaho.gov"),
    ("ID", "Ketchum", 2689, "Blaine", "clerk@ketchumidaho.org"),
    ("ID", "Hailey", 8348, "Blaine", "clerk@haileycityhall.org"),
    ("ID", "Sun Valley", 1406, "Blaine", "clerk@sunvalley.com"),
    ("ID", "Orofino", 3142, "Clearwater", "clerk@orofino-id.com"),
    ("ID", "Grangeville", 3141, "Idaho", "clerk@grangeville.us"),
    ("ID", "Weiser", 5507, "Washington", "clerk@cityofweiser.net"),
    ("ID", "Payette", 7655, "Payette", "clerk@cityofpayette.com"),
    ("ID", "Emmett", 6557, "Gem", "clerk@cityofemmett.org"),
    ("ID", "Homedale", 2633, "Owyhee", "clerk@cityofhomedale.com"),
    ("ID", "Rupert", 5554, "Minidoka", "clerk@rupert.id.gov"),
    ("ID", "Buhl", 4122, "Twin Falls", "clerk@cityofbuhl.us"),
    ("ID", "Gooding", 3567, "Gooding", "clerk@goodingidaho.org"),
    ("ID", "Shoshone", 1491, "Lincoln", "clerk@shoshonecity.com"),
    ("ID", "Challis", 1081, "Custer", "clerk@challisidaho.us"),
]

# Combine B-list
all_b_list = b_list_mt + b_list_me + b_list_wy + b_list_id

for row, data in enumerate(all_b_list, 2):
    for col, value in enumerate(data, 1):
        ws2.cell(row=row, column=col, value=value)
    ws2.cell(row=row, column=6, value="Not Sent")

# Set column widths
col_widths_b = [5, 20, 10, 15, 35, 12, 12, 8, 8, 8, 40]
for i, width in enumerate(col_widths_b, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ============================================================================
# SHEET 3: EMAIL TEMPLATES
# ============================================================================
ws3 = wb.create_sheet("Email Templates")

templates = [
    ("A-LIST TEMPLATE (Personalized)", """
Subject: Saw {town}'s permit forms - built something that might help

Hi {name},

I noticed {town} still uses paper permit applications [or: PDF forms on your website].

I'm Ian with Ito Geospatial - I built a simple permit tracking system specifically 
for Montana towns under 5,000 people. It's what iWorQ or OpenGov do, but at 1/10th 
the price ($300/month).

Here's a 30-second demo: [LINK]

The map already has {county} County parcel data loaded. You could be live in a week.

Would a quick call be useful? I'm based in Bozeman and happy to drive over.

Ian
Ito Geospatial
(406) 555-0123
"""),
    ("B-LIST TEMPLATE (Semi-personalized)", """
Subject: Permit tracking for {town} - $300/month

Hi,

Quick question: Is {town} still tracking permits on paper or spreadsheets?

I built a simple permit tracking system for small towns:
- Public can check permit status online (no more phone calls)
- Staff dashboard shows what needs review
- Map shows all active development
- Instant reports for council meetings

It's $300/month - no setup fees, cancel anytime.

30-second demo: [LINK]

Worth a look?

Ian
Ito Geospatial
"""),
    ("C-LIST TEMPLATE (Mass)", """
Subject: Stop losing track of permits

Hi,

I built a $300/month permit tracking system for small towns.

- Residents check status online
- Staff see everything on a map  
- Reports ready for council

Demo: [LINK]

Ian
Ito Geospatial
"""),
    ("FOLLOW-UP #1 (3 days later)", """
Subject: RE: Permit tracking for {town}

Hi {name},

Just bumping this up - did you get a chance to look at the demo?

Happy to answer any questions or jump on a quick call.

Ian
"""),
    ("FOLLOW-UP #2 (7 days later)", """
Subject: Last try - permit tracking

Hi {name},

I'll keep this short - if {town} ever needs an affordable way to track permits,
I'm here. The demo link still works: [LINK]

No hard feelings if it's not a fit. Best of luck!

Ian
Ito Geospatial
"""),
]

ws3['A1'] = "EMAIL TEMPLATES FOR OUTREACH"
ws3['A1'].font = Font(bold=True, size=14)
ws3.merge_cells('A1:D1')

row = 3
for title, template in templates:
    ws3.cell(row=row, column=1, value=title)
    ws3.cell(row=row, column=1).font = Font(bold=True, color="22c55e")
    row += 1
    for line in template.strip().split('\n'):
        ws3.cell(row=row, column=1, value=line)
        row += 1
    row += 2

ws3.column_dimensions['A'].width = 100

# ============================================================================
# SHEET 4: TRACKING DASHBOARD
# ============================================================================
ws4 = wb.create_sheet("Dashboard")

# Summary stats with formulas
ws4['A1'] = "ITO GEOSPATIAL - OUTREACH DASHBOARD"
ws4['A1'].font = Font(bold=True, size=16, color="22c55e")
ws4.merge_cells('A1:D1')

ws4['A3'] = "Campaign Statistics"
ws4['A3'].font = Font(bold=True, size=12)

stats = [
    ("Total A-List Prospects", f"={len(all_a_list)}"),
    ("Total B-List Prospects", f"={len(all_b_list)}"),
    ("Total Prospects", f"={len(all_a_list) + len(all_b_list)}"),
    ("", ""),
    ("A-List Contacted", '=COUNTIF(\'A-List (50 Priority)\'!N:N,"Contacted")'),
    ("A-List Responses", '=COUNTIF(\'A-List (50 Priority)\'!N:N,"Responded")'),
    ("A-List Demos Scheduled", '=COUNTIF(\'A-List (50 Priority)\'!N:N,"Demo Scheduled")'),
    ("A-List Won", '=COUNTIF(\'A-List (50 Priority)\'!N:N,"Won")'),
    ("", ""),
    ("B-List Sent", '=COUNTIF(\'B-List (200)\'!F:F,"Sent")'),
    ("B-List Opened", '=COUNTIF(\'B-List (200)\'!H:H,"Yes")'),
    ("B-List Clicked", '=COUNTIF(\'B-List (200)\'!I:I,"Yes")'),
    ("B-List Replied", '=COUNTIF(\'B-List (200)\'!J:J,"Yes")'),
]

for i, (label, formula) in enumerate(stats, 5):
    ws4.cell(row=i, column=1, value=label)
    ws4.cell(row=i, column=2, value=formula if formula else "")

# Conversion funnel
ws4['A20'] = "Conversion Targets"
ws4['A20'].font = Font(bold=True, size=12)

targets = [
    ("Metric", "Target", "Realistic", "Stretch"),
    ("Emails Sent", "250", "250", "500"),
    ("Opens (15-25%)", "38-63", "50", "100"),
    ("Clicks (2-5%)", "5-13", "8", "20"),
    ("Replies (0.5-2%)", "1-5", "3", "10"),
    ("Demos Scheduled", "2-5", "3", "8"),
    ("Deals Closed", "1-2", "1", "3"),
    ("Monthly Revenue", "$300-600", "$300", "$900"),
    ("Annual Revenue", "$3,600-7,200", "$3,600", "$10,800"),
]

for i, row_data in enumerate(targets, 21):
    for j, value in enumerate(row_data, 1):
        cell = ws4.cell(row=i, column=j, value=value)
        if i == 21:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

# Timeline
ws4['A32'] = "Outreach Timeline"
ws4['A32'].font = Font(bold=True, size=12)

timeline = [
    ("Week", "Task", "Target"),
    ("Week 1", "Finalize demo, build list", "Product ready"),
    ("Week 1", "A-List research & personalization", "50 personalized emails drafted"),
    ("Week 2", "Send A-List emails (10/day)", "50 sent"),
    ("Week 2", "Follow-up #1 on early sends", "Warm leads identified"),
    ("Week 3", "Send B-List batch 1 (100)", "100 sent"),
    ("Week 3", "A-List follow-up #2", "Demo conversations"),
    ("Week 4", "Send B-List batch 2 (100)", "200 total sent"),
    ("Week 4", "Demo calls, close deals", "1-2 pilots signed"),
]

for i, row_data in enumerate(timeline, 33):
    for j, value in enumerate(row_data, 1):
        cell = ws4.cell(row=i, column=j, value=value)
        if i == 33:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

for col in range(1, 5):
    ws4.column_dimensions[get_column_letter(col)].width = 25

# ============================================================================
# SHEET 5: COMPETITOR NOTES
# ============================================================================
ws5 = wb.create_sheet("Competitor Intel")

ws5['A1'] = "COMPETITOR INTELLIGENCE"
ws5['A1'].font = Font(bold=True, size=14)

competitors = [
    ("Competitor", "Price Range", "Target Market", "Weaknesses", "Our Advantage"),
    ("iWorQ", "$3,500-35,000/yr", "2,000-75,000 pop", "Mixed reviews, no GIS", "Half price, GIS built-in"),
    ("Cloudpermit", "$10,000-25,000/yr", "Any size", "Canadian, implementation issues", "US-focused, faster setup"),
    ("OpenGov", "$40,000-80,000/yr", "5,000-100,000 pop", "Expensive, enterprise sales", "10x cheaper, same features"),
    ("GovPilot", "$5,000-50,000/yr", "5,000-50,000 pop", "Complex, feature bloat", "Simple, focused"),
    ("ClearForms", "Unknown (low)", "500-4M pop", "Forms tool, not permit system", "Full workflow"),
    ("BasicGov", "$1,500/seat/yr", "Any size", "Per-seat pricing adds up", "Unlimited users"),
    ("Accela", "$100,000-500,000+/yr", "Large cities", "Enterprise only", "We don't compete here"),
    ("Tyler/EnerGov", "$75,000+/yr", "Large cities", "Enterprise only", "We don't compete here"),
]

for i, row_data in enumerate(competitors, 3):
    for j, value in enumerate(row_data, 1):
        cell = ws5.cell(row=i, column=j, value=value)
        if i == 3:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

for col in range(1, 6):
    ws5.column_dimensions[get_column_letter(col)].width = 30

# ============================================================================
# SAVE
# ============================================================================
output_path = '/home/claude/ito-permits-v2/ito_sales_pipeline.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
print(f"A-List: {len(all_a_list)} prospects")
print(f"B-List: {len(all_b_list)} prospects")
print(f"Total: {len(all_a_list) + len(all_b_list)} prospects")
